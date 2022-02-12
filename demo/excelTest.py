#必要なモジュールのインポート
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

#Excelファイルをwbという変数に読み込む
wb = openpyxl.load_workbook("D:\RPAtest.xlsx")

#Excelファイルのシート名「Sheet1」をwsという変数に読み込む
ws = wb["Sheet1"]

#ブラウザを起動し、Google検索ページを開く。念のため待機処理を入れる
#このコードでバージョンの問題から生じるエラーを回避できている
driver = webdriver.Chrome(ChromeDriverManager().install())#Chromeを起動
driver.get("https://www.google.com/?hl=ja")#Google検索ページを開く
WebDriverWait(driver, 15).until(EC.visibility_of_all_elements_located)#待機処理

#奈良、京都、大阪、滋賀とそれぞれ入力するため、for文で繰り返し処理
for i in range(3,7):#range(3,7)は3,4,5,6（ExcelのB3,B4,B5,B6セルに対応）
    element = driver.find_element_by_name("q")#検索窓要素を取得。name以外にもid、css_selectorなど
    keyword = ws.cell(row=i, column=2).value#ExcelのセルBiの値をkeywordという変数に入力。rowは行、columnは列
    element.send_keys(keyword)#.send_keys()で検索窓にkeywordを入力
    element.submit()#検索を実行
    WebDriverWait(driver, 15).until(EC.visibility_of_all_elements_located)#待機処理
    titleElements = driver.find_elements_by_class_name("LC20lb")#検索結果のタイトル要素を取得
    output = titleElements[0].text#最初の検索結果のタイトルテキストをoutputに入力
    ws.cell(row=i, column=3).value = output#検索結果のタイトルをExcelのセルCiに入力
    driver.back()#ブラウザの「戻る」を実行。繰り返しはここまで。
wb.save("D:\RPAtest.xlsx") #Excelを保存